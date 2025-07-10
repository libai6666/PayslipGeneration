"""
数据导入模块
用于导入员工数据和导出模板
"""

import os
import csv
from openpyxl import Workbook, load_workbook
from datetime import datetime


def import_employee_data(file_path):
    """
    导入员工数据 - 不依赖pandas，支持中文表头，自动检测表头行，支持多表头格式
    
    参数:
        file_path (str): 数据文件路径
    
    返回:
        list: 员工数据字典列表
    """
    # 获取文件扩展名
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    employees = []
    
    # 列名映射（中文 -> 英文）
    column_map = {
        '姓名': 'name',
        '月份': 'month',
        '基本工资': 'base_salary',
        '应出勤天数': 'required_days',
        '实际出勤天数': 'actual_days',
        '夜班补助': 'night_shift',
        '高温补贴': 'high_temp',
        '迟到罚款': 'late_fine',
        '其他': 'others'
    }
    
    # 反向映射字典，用于错误提示（英文 -> 中文）
    reverse_map = {v: k for k, v in column_map.items()}
    
    # 必要的中文列
    required_cn_columns = ['姓名', '基本工资', '应出勤天数', '实际出勤天数']
    
    try:
        if ext in ['.xlsx', '.xls']:
            # 使用openpyxl读取Excel文件
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # 按行扫描工作表，识别表头行和数据行
            current_row = 1
            max_row = ws.max_row
            
            while current_row <= max_row:
                # 尝试将当前行作为表头
                headers = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=current_row, column=col).value
                    if cell_value is not None:
                        headers.append(str(cell_value).strip())
                    else:
                        headers.append(None)
                
                # 检查是否为表头行
                found_headers = [h for h in headers if h is not None]
                missing_count = 0
                for required_cn in required_cn_columns:
                    if not any(required_cn in h or h == required_cn for h in found_headers):
                        missing_count += 1
                
                if missing_count == 0:  # 找到了可能的表头行
                    print(f"在第{current_row}行检测到表头: {headers}")
                    
                    # 映射列名
                    mapped_indices = {}
                    for i, header in enumerate(headers):
                        if header is None:
                            continue
                            
                        header_lower = header.lower()
                        
                        # 1. 首先尝试精确匹配
                        if header in column_map:
                            mapped_indices[column_map[header]] = i
                            continue
                        
                        # 2. 然后尝试不区分大小写的精确匹配
                        for cn, en in column_map.items():
                            if cn.lower() == header_lower:
                                mapped_indices[en] = i
                                break
                                
                        # 3. 最后尝试包含关系匹配
                        if not any(cn.lower() == header_lower for cn in column_map):
                            for cn, en in column_map.items():
                                if cn.lower() in header_lower or header_lower in cn.lower():
                                    mapped_indices[en] = i
                                    break
                    
                    print(f"列映射结果: {mapped_indices}")
                    
                    # 检查必要的列是否存在
                    required_columns = ['name', 'base_salary', 'required_days', 'actual_days']
                    missing_columns = [col for col in required_columns if col not in mapped_indices]
                    if missing_columns:
                        print(f"第{current_row}行缺少必要的列：{missing_columns}，尝试下一行")
                        current_row += 1
                        continue
                    
                    # 读取下一行作为数据
                    if current_row + 1 <= max_row:
                        data_row = current_row + 1
                        employee = {}
                        
                        # 获取映射后的值
                        for en, i in mapped_indices.items():
                            cell_value = ws.cell(row=data_row, column=i+1).value
                            if cell_value is not None:
                                employee[en] = cell_value
                        
                        # 如果有姓名，添加到员工列表
                        if employee.get('name'):
                            # 确保数值字段类型正确
                            for field in ['base_salary', 'required_days', 'actual_days', 
                                         'night_shift', 'high_temp', 'late_fine', 'others']:
                                if field in employee:
                                    try:
                                        if field in ['required_days', 'actual_days']:
                                            employee[field] = int(float(str(employee[field]).replace(',', '')))
                                        else:
                                            employee[field] = float(str(employee[field]).replace(',', ''))
                                    except (ValueError, TypeError):
                                        print(f"警告：{field}字段的值'{employee[field]}'无法转换为数值，已设为0")
                                        employee[field] = 0
                                else:
                                    employee[field] = 0
                            
                            # 处理月份字段
                            if 'month' in employee:
                                try:
                                    employee['month'] = int(float(str(employee['month']).replace(',', '')))
                                except (ValueError, TypeError):
                                    employee['month'] = datetime.now().month
                                    print(f"警告：月份字段值无效，已设为当前月份{employee['month']}")
                            else:
                                employee['month'] = datetime.now().month
                                print(f"警告：缺少月份字段，已设为当前月份{employee['month']}")
                            
                            employees.append(employee)
                            print(f"成功导入员工：{employee['name']}")
                        
                        # 跳过当前表头和数据行，以及可能的空行
                        current_row = data_row + 1
                        # 跳过空行
                        while current_row <= max_row:
                            has_data = False
                            for col in range(1, ws.max_column + 1):
                                if ws.cell(row=current_row, column=col).value is not None:
                                    has_data = True
                                    break
                            if has_data:
                                break
                            current_row += 1
                    else:
                        # 已经到达文件末尾
                        break
                else:
                    # 不是表头行，继续检查下一行
                    current_row += 1
            
            wb.close()
            
        elif ext == '.csv':
            # 使用csv模块读取CSV文件
            with open(file_path, newline='', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = list(reader)
                
                if not rows:
                    raise ValueError("CSV文件为空")
                
                # 按行扫描，识别表头行和数据行
                row_index = 0
                while row_index < len(rows):
                    # 尝试将当前行作为表头
                    headers = [h.strip() for h in rows[row_index] if h]
                    
                    # 检查是否为表头行
                    missing_count = 0
                    for required_cn in required_cn_columns:
                        if not any(required_cn in h or h == required_cn for h in headers):
                            missing_count += 1
                    
                    if missing_count == 0:  # 找到了可能的表头行
                        print(f"在第{row_index+1}行检测到表头: {headers}")
                        
                        # 映射列名
                        mapped_indices = {}
                        for i, header in enumerate(headers):
                            if not header:
                                continue
                                
                            header_lower = header.lower()
                            
                            # 1. 首先尝试精确匹配
                            if header in column_map:
                                mapped_indices[column_map[header]] = i
                                continue
                            
                            # 2. 然后尝试不区分大小写的精确匹配
                            for cn, en in column_map.items():
                                if cn.lower() == header_lower:
                                    mapped_indices[en] = i
                                    break
                                    
                            # 3. 最后尝试包含关系匹配
                            if not any(cn.lower() == header_lower for cn in column_map):
                                for cn, en in column_map.items():
                                    if cn.lower() in header_lower or header_lower in cn.lower():
                                        mapped_indices[en] = i
                                        break
                        
                        print(f"列映射结果: {mapped_indices}")
                        
                        # 检查必要的列是否存在
                        required_columns = ['name', 'base_salary', 'required_days', 'actual_days']
                        missing_columns = [col for col in required_columns if col not in mapped_indices]
                        if missing_columns:
                            print(f"第{row_index+1}行缺少必要的列：{missing_columns}，尝试下一行")
                            row_index += 1
                            continue
                        
                        # 读取下一行作为数据
                        if row_index + 1 < len(rows):
                            data_row = row_index + 1
                            row_data = rows[data_row]
                            employee = {}
                            
                            # 获取映射后的值
                            for en, i in mapped_indices.items():
                                if i < len(row_data):
                                    cell_value = row_data[i].strip()
                                    if cell_value:
                                        employee[en] = cell_value
                            
                            # 如果有姓名，添加到员工列表
                            if employee.get('name'):
                                # 确保数值字段类型正确
                                for field in ['base_salary', 'required_days', 'actual_days', 
                                             'night_shift', 'high_temp', 'late_fine', 'others']:
                                    if field in employee:
                                        try:
                                            if field in ['required_days', 'actual_days']:
                                                employee[field] = int(float(str(employee[field]).replace(',', '')))
                                            else:
                                                employee[field] = float(str(employee[field]).replace(',', ''))
                                        except (ValueError, TypeError):
                                            print(f"警告：{field}字段的值'{employee[field]}'无法转换为数值，已设为0")
                                            employee[field] = 0
                                    else:
                                        employee[field] = 0
                                
                                # 处理月份字段
                                if 'month' in employee:
                                    try:
                                        employee['month'] = int(float(str(employee['month']).replace(',', '')))
                                    except (ValueError, TypeError):
                                        employee['month'] = datetime.now().month
                                        print(f"警告：月份字段值无效，已设为当前月份{employee['month']}")
                                else:
                                    employee['month'] = datetime.now().month
                                    print(f"警告：缺少月份字段，已设为当前月份{employee['month']}")
                                
                                employees.append(employee)
                                print(f"成功导入员工：{employee['name']}")
                            
                            # 跳过当前表头和数据行，以及可能的空行
                            row_index = data_row + 1
                            # 跳过空行
                            while row_index < len(rows):
                                if any(cell.strip() for cell in rows[row_index] if cell):
                                    break
                                row_index += 1
                        else:
                            # 已经到达文件末尾
                            break
                    else:
                        # 不是表头行，继续检查下一行
                        row_index += 1
        else:
            raise ValueError(f"不支持的文件类型：{ext}")
        
        if not employees:
            raise ValueError("没有找到有效的员工数据")
        
        return employees
    
    except Exception as e:
        print(f"导入数据时出错：{str(e)}")
        raise ValueError(f"导入数据时出错：{str(e)}")


def export_template(file_path):
    """
    导出数据导入模板 - 不依赖pandas，使用多表头格式
    
    参数:
        file_path (str): 模板文件保存路径
    """
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "员工数据"
        
        # 表头内容
        headers = ["姓名", "月份", "基本工资", "应出勤天数", "实际出勤天数", 
                  "夜班补助", "高温补贴", "迟到罚款", "其他"]
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # 当前行
        current_row = 1
        
        # 示例1
        # 添加表头
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col).value = header
        
        # 添加示例数据1
        current_row += 1
        ws.cell(row=current_row, column=1).value = "示例：张三"
        ws.cell(row=current_row, column=2).value = datetime.now().month  # 当前月份作为示例
        ws.cell(row=current_row, column=3).value = 5000
        ws.cell(row=current_row, column=4).value = 30
        ws.cell(row=current_row, column=5).value = 28
        ws.cell(row=current_row, column=6).value = 200
        ws.cell(row=current_row, column=7).value = 100
        ws.cell(row=current_row, column=8).value = -50
        ws.cell(row=current_row, column=9).value = 0
        
        # 添加空行
        current_row += 2
        
        # 示例2
        # 添加表头
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col).value = header
        
        # 添加示例数据2
        current_row += 1
        ws.cell(row=current_row, column=1).value = "示例：李四"
        ws.cell(row=current_row, column=2).value = datetime.now().month  # 当前月份作为示例
        ws.cell(row=current_row, column=3).value = 8000
        ws.cell(row=current_row, column=4).value = 30
        ws.cell(row=current_row, column=5).value = 29
        ws.cell(row=current_row, column=6).value = 500
        ws.cell(row=current_row, column=7).value = 500
        ws.cell(row=current_row, column=8).value = -1000
        ws.cell(row=current_row, column=9).value = 0
        
        # 添加空行
        current_row += 2
        
        # 示例3 - 用于用户填写
        # 添加表头
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col).value = header
        
        # 保存文件
        wb.save(file_path)
        return True
    except Exception as e:
        print(f"导出模板时出错：{str(e)}")
        raise Exception(f"导出模板时出错：{str(e)}")