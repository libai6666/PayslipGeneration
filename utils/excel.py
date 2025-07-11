"""
Excel工资条生成模块
用于生成工资条Excel文件
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def generate_excel(employee_data, output_path=None):
    """
    生成工资条Excel文件
    
    参数:
        employee_data (dict): 员工工资数据
        output_path (str, optional): 输出文件路径，默认为当前目录
    
    返回:
        str: 生成的Excel文件路径
    """
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "工资条"
    
    # 获取年份和月份
    year = employee_data.get('year', datetime.now().year)
    month = employee_data.get('month', datetime.now().month)
    
    # 设置列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 15
    
    # 添加表头（直接从第1行开始）
    row = 1
    ws.cell(row=row, column=1).value = "姓名"
    ws.cell(row=row, column=2).value = "年份"
    ws.cell(row=row, column=3).value = "月份"
    ws.cell(row=row, column=4).value = "基本工资"
    ws.cell(row=row, column=5).value = "应出勤天数"
    ws.cell(row=row, column=6).value = "实际出勤天数"
    ws.cell(row=row, column=7).value = "夜班补助"
    ws.cell(row=row, column=8).value = "高温补贴"
    ws.cell(row=row, column=9).value = "迟到罚款"
    ws.cell(row=row, column=10).value = "其他"
    ws.cell(row=row, column=11).value = "缺勤扣款"
    ws.cell(row=row, column=12).value = "实发工资"
    ws.cell(row=row, column=13).value = "签字"
    
    # 设置标题样式
    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        set_cell_style(cell, 'header')
    
    # 添加数据（从第2行开始）
    row = 2
    ws.cell(row=row, column=1).value = employee_data.get('name', '')
    ws.cell(row=row, column=2).value = year
    ws.cell(row=row, column=3).value = month
    ws.cell(row=row, column=4).value = employee_data.get('base_salary', 0)
    ws.cell(row=row, column=5).value = employee_data.get('required_days', 0)
    ws.cell(row=row, column=6).value = employee_data.get('actual_days', 0)
    ws.cell(row=row, column=7).value = employee_data.get('night_shift', 0)
    ws.cell(row=row, column=8).value = employee_data.get('high_temp', 0)
    ws.cell(row=row, column=9).value = employee_data.get('late_fine', 0)
    ws.cell(row=row, column=10).value = employee_data.get('others', 0)
    ws.cell(row=row, column=11).value = employee_data.get('absence_deduction', 0)
    ws.cell(row=row, column=12).value = employee_data.get('net_salary', 0)
    ws.cell(row=row, column=13).value = employee_data.get('signature', '')
    
    # 设置数据样式
    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        set_cell_style(cell, 'normal')
    
    # 设置特殊单元格样式
    set_cell_style(ws.cell(row=row, column=11), 'deduction')
    set_cell_style(ws.cell(row=row, column=12), 'total')
    
    # 确定保存路径
    if not output_path:
        # 默认保存到桌面
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"{year}年{month}月_工资条_{employee_data.get('name', '')}_{timestamp}.xlsx"
        output_path = os.path.join(desktop, filename)
    
    # 保存工作簿
    wb.save(output_path)
    return output_path


def batch_generate_excel(employees, output_dir=None):
    """
    批量生成工资条Excel文件
    
    参数:
        employees (list): 员工数据字典列表
        output_dir (str, optional): 输出目录，默认为桌面
    
    返回:
        list: 生成的Excel文件路径列表
    """
    if output_dir is None:
        # 默认保存到桌面
        output_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
    
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    
    # 时间戳（用于生成文件名）
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    
    # 生成的文件路径列表
    file_paths = []
    
    # 为每位员工生成工资条
    for employee in employees:
        # 获取年份和月份
        year = employee.get('year', datetime.now().year)
        month = employee.get('month', datetime.now().month)
        
        # 生成文件名
        filename = f"{year}年{month}月_工资条_{employee.get('name', 'unknown')}_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, filename)
        
        # 调用单个生成函数
        generate_excel(employee, output_path)
        file_paths.append(output_path)
    
    return file_paths


def generate_summary_excel(employees, month=None, output_path=None):
    """
    生成汇总工资条Excel文件 - 每个员工数据前都有表头
    
    参数:
        employees (list): 员工数据字典列表
        month (int, optional): 默认月份，默认为当前月份
        output_path (str, optional): 输出文件路径，默认为桌面
    
    返回:
        str: 生成的Excel文件路径
    """
    if not employees:
        raise ValueError("没有员工数据")
    
    # 默认使用当前月份和年份
    year = datetime.now().year
    if month is None:
        month = datetime.now().month
    
    # 尝试从员工数据中获取年份
    if employees and 'year' in employees[0]:
        year = employees[0]['year']
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "工资表"
    
    # 设置列宽
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    for col in columns:
        ws.column_dimensions[col].width = 15
    
    # 表头内容
    headers = ["姓名", "年份", "月份", "基本工资", "应出勤天数", "实际出勤天数", 
              "夜班补助", "高温补贴", "迟到罚款", "其他", "缺勤扣款", "实发工资", "签字"]
    
    # 当前行
    current_row = 1
    
    # 为每个员工添加表头和数据
    for employee in employees:
        # 添加表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            set_cell_style(cell, 'header')
        
        # 添加员工数据
        current_row += 1
        ws.cell(row=current_row, column=1).value = employee.get('name', '')
        ws.cell(row=current_row, column=2).value = employee.get('year', year)
        ws.cell(row=current_row, column=3).value = employee.get('month', month)
        ws.cell(row=current_row, column=4).value = employee.get('base_salary', 0)
        ws.cell(row=current_row, column=5).value = employee.get('required_days', 0)
        ws.cell(row=current_row, column=6).value = employee.get('actual_days', 0)
        ws.cell(row=current_row, column=7).value = employee.get('night_shift', 0)
        ws.cell(row=current_row, column=8).value = employee.get('high_temp', 0)
        ws.cell(row=current_row, column=9).value = employee.get('late_fine', 0)
        ws.cell(row=current_row, column=10).value = employee.get('others', 0)
        ws.cell(row=current_row, column=11).value = employee.get('absence_deduction', 0)
        ws.cell(row=current_row, column=12).value = employee.get('net_salary', 0)
        ws.cell(row=current_row, column=13).value = employee.get('signature', '')
        
        # 设置样式
        for col in range(1, 14):
            cell = ws.cell(row=current_row, column=col)
            set_cell_style(cell, 'normal')
        
        # 设置特殊单元格样式
        set_cell_style(ws.cell(row=current_row, column=11), 'deduction')
        set_cell_style(ws.cell(row=current_row, column=12), 'total')
        
        # 添加空行（除非是最后一个员工）
        if employee != employees[-1]:
            current_row += 2  # 增加2行，留出一个空行
    
    # 计算总计并添加到最后一行
    if employees:
        # 空两行
        current_row += 2
        
        # 添加总计行
        ws.cell(row=current_row, column=1).value = "总计"
        set_cell_style(ws.cell(row=current_row, column=1), 'header')
        
        # 计算各列总计
        total_base_salary = sum(e.get('base_salary', 0) for e in employees)
        total_high_temp = sum(e.get('high_temp', 0) for e in employees)
        total_absence_deduction = sum(e.get('absence_deduction', 0) for e in employees)
        total_net_salary = sum(e.get('net_salary', 0) for e in employees)
        
        # 添加总计值
        ws.cell(row=current_row, column=4).value = total_base_salary
        ws.cell(row=current_row, column=8).value = total_high_temp
        ws.cell(row=current_row, column=11).value = total_absence_deduction
        ws.cell(row=current_row, column=12).value = total_net_salary
        
        # 设置总计行样式
        for col in [4, 8, 11, 12]:
            set_cell_style(ws.cell(row=current_row, column=col), 'total')
    
    # 确定保存路径
    if not output_path:
        # 默认保存到桌面
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        filename = f"{year}年{month}月_工资表_{timestamp}.xlsx"
        output_path = os.path.join(desktop, filename)
    
    # 保存工作簿
    wb.save(output_path)
    return output_path


def set_cell_style(cell, style_type):
    """
    设置单元格样式
    
    参数:
        cell (openpyxl.cell): 单元格对象
        style_type (str): 样式类型，如'header', 'normal', 'deduction', 'total'
    """
    # 基本样式 - 所有单元格共有
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 特定样式
    if style_type == 'header':
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    elif style_type == 'normal':
        cell.font = Font(size=11)
    
    elif style_type == 'deduction':
        cell.font = Font(size=11, color="FF0000")  # 红色字体表示扣款
    
    elif style_type == 'total':
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")